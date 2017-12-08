import openpyxl
import xlrd
import string


def open_workbook(wb_name):
    wb = openpyxl.load_workbook(wb_name)
    return wb


def open_worksheet(workbook, sheet):
    workbook.get_sheet_by_name(sheet)


def load_name_ranges(wb, range_name):
    workbook = openpyxl.load_workbook(wb)
    return workbook.get_named_ranges()

    #Name = workbook.name_map[range_name][0]
    #Sheet, rowxlo, rowxhi, colxlo, colxhi = Name.area2d()
    #for i in range(rowxhi):
    #    print(Sheet.cell(i, 0))


if __name__ == '__main__':
    #xl = win32com.client.Dispatch('excel.application')
    lib = 'ReOp_Library_v2_it2.xlsx'

    named_reference = {}

    # This will load all of the named references into a dictionary for later use

    workbook = xlrd.open_workbook(lib)
    for (name, scope), v in workbook.name_and_scope_map.iteritems():
        cell = workbook.name_and_scope_map.get((name, scope))
        (sheetName, ref) = cell.formula_text.split('!')
        print name, cell.formula_text

        named_reference[name] = cell.formula_text
