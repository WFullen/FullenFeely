
import xlsxwriter
from openpyxl import load_workbook

# Upon login, create an initial workbook and begin to ask a questionnaire

def define_initial_workbook():
    file_name = raw_input("Give a name for your Decision Support Model (DSM):")
    version_number = raw_input("Give a version number for your DSM:")



    libWorkbook = xlsxwriter.Workbook(file_name + '_Library_' + version_number)
    modelWorkbook = xlsxwriter.Workbook(file_name + '_Model_' + version_number)


def create_library_sheets(libraryWorkbook):

    # Function creates sheets on the library workbook (at least one), and prompts the
    # user for any additional sheets. The sheets will be stored in an array for access later
    # and will be discarded once the spreadsheet has been populated.

    user_response = 'yes'
    number_of_sheets = 1
    sheet_names = []
    while user_response == 'yes':
        sheet_name = raw_input("Please provide a name for the option you want in your DSM:")
        worksheet = libraryWorkbook.add_worksheet(sheet_name)
        sheet_names.append(sheet_name)
        number_of_sheets += number_of_sheets + 1
        user_response = raw_input("Would you like to enter another Options tab? (please enter yes/no)")


    # This method will be used to populate the Library sheets once all have been defined and named.
    # This method is called by createLibrarySheets

