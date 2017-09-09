
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl import *

class Library(object):

    def __init__(self, name, choices=[]):
        self.choices = choices
        self.name = name

    def instantiate_workbook(self):
        # Check if the workbook exists
        if os.path.isfile(self.name):

            wb = load_workbook(self.name)
        else:
            wb = Workbook()
            file_save = raw_input("Where would you like to save your file? (Please provide the full path)")
            wb.save(filename=file_save)

        return wb

    def update_choices(self):
        choice_list = raw_input("Please supply a list of choices for comparison (list of variables separated by commas)")
        self.choices = choice_list.split(",")

    def update_metrics(src, dest, name_of_sheet):
        # Open an xlsx for reading
        wb = load_workbook(filename=src)

        # Get the sheet by name
        ws = wb.get_sheet_by_name(name_of_sheet)

        line_num = 3
        add_more_cats = 'Y'
        while add_more_cats == 'Y':


            # Must provide at least one Category
            category = raw_input("Please provide a category for sheet" + name_of_sheet)
            ws['B' + line_num] = category

            line_num += 2

            add_more_subcats = 'Y'

            while add_more_subcats == 'Y':
                sub_category = raw_input("Please provide a sub-category for this category" + name_of_sheet)
                ws['B' + line_num] = sub_category

                line_num += 1

                add_more_metrics = 'Y'
                while add_more_metrics == 'Y':
                    metric = raw_input("Please provide a metric for this sub-category" + name_of_sheet)
                    ws['B' + line_num] = metric
                    line_num += 1

                    add_more_metrics = raw_input("Would you like to add an additional metric? (Y/N)")

                add_more_subcats = raw_input("Would you like to add an additional metric? (Y/N)")

            add_more_cats = raw_input("Would you like to add an additional metric? (Y/N)")

        # save the csv file
        wb.save(dest)